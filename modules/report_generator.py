import pandas as pd
from openpyxl.styles import PatternFill

def generate_report(data, output_path):
    """Genera un reporte en Excel con los datos auditados."""
    df = pd.DataFrame(data)
    
    # Estilo condicional para resaltar problemas
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="Auditoría")
    
    # Formato adicional
    workbook = writer.book
    worksheet = writer.sheets["Auditoría"]
    
    # Resaltar alto uso de CPU/RAM
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in ["cpu_usage", "ram_usage"]:
        col_idx = df.columns.get_loc(col)
        if isinstance(col_idx, int):
            excel_col = col_idx + 1  # Excel columns are 1-based
            for idx, value in enumerate(df[col]):
                if value > 80:  # Umbral de alerta
                    worksheet.cell(row=idx+2, column=excel_col).fill = red_fill
        else:
            # Handle slices or arrays if needed
            pass
    
    writer.close()